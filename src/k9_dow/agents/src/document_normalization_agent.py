# SPDX-License-Identifier: Apache-2.0
# DoW Architecture Workbench — DocumentNormalizationAgent (SBB)
#
# Pure extraction agent — no LLM invocation.
# Converts uploaded documents (MD, TXT, PDF, DOCX, XLSX, CSV) into
# normalized Markdown text with metadata.

import logging
from pathlib import Path
from typing import Any, Dict, Optional

from k9_aif_abb.k9_core.agent.base_agent import BaseAgent

log = logging.getLogger(__name__)


class DocumentNormalizationAgent(BaseAgent):
    layer = "DoW DocumentNormalization SBB"

    def __init__(self, config: Optional[Dict[str, Any]] = None, monitor=None, **kwargs):
        super().__init__(config or {}, monitor=monitor, **kwargs)

    def execute(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        warnings: list[str] = []
        markdown = ""
        metadata: Dict[str, Any] = {}

        source_markdown = payload.get("source_markdown") or ""
        raw_path = (payload.get("metadata") or {}).get("raw_path")

        if source_markdown:
            markdown = source_markdown
            metadata["conversion_method"] = "passthrough"
        elif raw_path:
            path = Path(raw_path)
            markdown, meta_warnings = self._extract_from_file(path)
            warnings.extend(meta_warnings)
            metadata["conversion_method"] = f"file_extraction:{path.suffix}"
            metadata["file_size"] = path.stat().st_size if path.exists() else 0
        else:
            self.publish_event({"type": "AgentCompleted", "agent": "DocumentNormalizationAgent", "status": "failed"})
            return {
                "agent": "DocumentNormalizationAgent",
                "output": "No source_markdown or raw_path provided",
                "status": "failed",
                "errors": ["No source_markdown or raw_path provided"],
            }

        if not markdown or not markdown.strip():
            self.publish_event({"type": "AgentCompleted", "agent": "DocumentNormalizationAgent", "status": "failed"})
            return {
                "agent": "DocumentNormalizationAgent",
                "output": "Document produced empty text after normalization",
                "status": "failed",
                "errors": ["Document produced empty text after normalization"],
                "warnings": warnings,
            }

        metadata["char_count"] = len(markdown)
        metadata["line_count"] = markdown.count("\n") + 1

        self.publish_event({"type": "AgentCompleted", "agent": "DocumentNormalizationAgent", "status": "completed"})
        return {
            "agent": "DocumentNormalizationAgent",
            "output": markdown,
            "metadata": metadata,
            "warnings": warnings,
        }

    # -- File extraction dispatch -----------------------------------------

    def _extract_from_file(self, path: Path) -> tuple[str, list[str]]:
        warnings: list[str] = []
        suffix = path.suffix.lower()

        if not path.exists():
            return "", [f"File not found: {path}"]

        if suffix in (".md", ".markdown", ".txt"):
            return path.read_text(encoding="utf-8", errors="ignore"), warnings

        if suffix == ".docx":
            return self._extract_docx(path, warnings)

        if suffix == ".pdf":
            return self._extract_pdf(path, warnings)

        if suffix in (".csv", ".xlsx"):
            return self._extract_tabular(path, suffix, warnings)

        warnings.append(f"Unsupported file type: {suffix}, attempting raw text decode")
        try:
            return path.read_text(encoding="utf-8", errors="ignore"), warnings
        except Exception as exc:
            return "", warnings + [f"Raw decode failed: {exc}"]

    # -- Format-specific extractors ---------------------------------------

    def _extract_docx(self, path: Path, warnings: list[str]) -> tuple[str, list[str]]:
        try:
            from docx import Document

            doc = Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs), warnings
        except ImportError:
            warnings.append("python-docx not installed; pip install python-docx")
            return "", warnings
        except Exception as exc:
            warnings.append(f"DOCX extraction failed: {exc}")
            return "", warnings

    def _extract_pdf(self, path: Path, warnings: list[str]) -> tuple[str, list[str]]:
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            pages = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(f"<!-- Page {i+1} -->\n{text}")
            if not pages:
                warnings.append("PDF produced no extractable text; may need OCR")
            return "\n\n".join(pages), warnings
        except ImportError:
            warnings.append("pypdf not installed; pip install pypdf")
            return "", warnings
        except Exception as exc:
            warnings.append(f"PDF extraction failed: {exc}")
            return "", warnings

    def _extract_tabular(
        self, path: Path, suffix: str, warnings: list[str]
    ) -> tuple[str, list[str]]:
        try:
            if suffix == ".csv":
                import csv

                with open(path, encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
            else:
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(str(path), read_only=True, data_only=True)
                    ws = wb.active
                    rows = [[str(c.value or "") for c in row] for row in ws.iter_rows()]
                    wb.close()
                except ImportError:
                    warnings.append("openpyxl not installed; pip install openpyxl")
                    return "", warnings

            if not rows:
                return "", warnings + ["Empty spreadsheet"]

            header = "| " + " | ".join(rows[0]) + " |"
            sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
            body = "\n".join("| " + " | ".join(row) + " |" for row in rows[1:])
            return f"{header}\n{sep}\n{body}", warnings

        except Exception as exc:
            warnings.append(f"Tabular extraction failed: {exc}")
            return "", warnings
